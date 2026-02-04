"""Report generation service for PDF and Excel exports."""
import io
from datetime import datetime, date
from typing import List, Dict, Any, Optional
from decimal import Decimal

# Excel export using openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# PDF export using reportlab
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


def generate_expense_excel(
    expenses: List[Dict[str, Any]],
    summary: Dict[str, Any],
    title: str = "Expense Report"
) -> bytes:
    """Generate Excel report for expenses."""
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Report"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Summary section
    ws['A3'] = "Summary"
    ws['A3'].font = Font(bold=True, size=12)
    
    ws['A4'] = "Total Amount:"
    ws['B4'] = summary.get('total_amount', 0)
    ws['A5'] = "Total Count:"
    ws['B5'] = summary.get('total_count', 0)
    ws['A6'] = "Date Range:"
    ws['B6'] = summary.get('date_range', 'N/A')
    
    # Headers
    headers = ['ID', 'Title', 'User', 'Department', 'Project', 'Amount', 'Status', 'Submitted Date']
    header_row = 9
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row_idx, expense in enumerate(expenses, header_row + 1):
        data = [
            expense.get('expense_id', ''),
            expense.get('expense_title', ''),
            expense.get('user_name', ''),
            expense.get('department_name', ''),
            expense.get('project_name', ''),
            expense.get('total_amount', 0),
            expense.get('status', ''),
            expense.get('submitted_at', '')
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if col == 6:  # Amount column
                cell.number_format = '#,##0.00'
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_expense_pdf(
    expenses: List[Dict[str, Any]],
    summary: Dict[str, Any],
    title: str = "Expense Report"
) -> bytes:
    """Generate PDF report for expenses."""
    if not PDF_AVAILABLE:
        raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20,
        alignment=1  # Center
    )
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 10))
    
    # Summary
    summary_data = [
        ['Total Amount:', f"{summary.get('total_amount', 0):,.2f}"],
        ['Total Count:', str(summary.get('total_count', 0))],
        ['Date Range:', summary.get('date_range', 'N/A')],
        ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M')]
    ]
    
    summary_table = Table(summary_data, colWidths=[100, 200])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Expense table
    if expenses:
        headers = ['ID', 'Title', 'User', 'Dept', 'Amount', 'Status', 'Date']
        table_data = [headers]
        
        for expense in expenses:
            row = [
                expense.get('expense_id', '')[:8] + '...',
                expense.get('expense_title', '')[:20],
                expense.get('user_name', '')[:15],
                (expense.get('department_name', '') or '-')[:10],
                f"{expense.get('total_amount', 0):,.2f}",
                expense.get('status', ''),
                str(expense.get('submitted_at', '-'))[:10]
            ]
            table_data.append(row)
        
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.27, 0.45, 0.77)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("No expenses found for the specified criteria.", styles['Normal']))
    
    doc.build(elements)
    output.seek(0)
    return output.getvalue()


def generate_tax_report_excel(
    items: List[Dict[str, Any]],
    year: int
) -> bytes:
    """Generate tax report in Excel format."""
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export")
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Tax Report {year}"
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = f"Tax Report - {year}"
    ws['A1'].font = Font(bold=True, size=16)
    
    # Headers
    headers = ['Date', 'Vendor', 'Description', 'Category', 'Amount', 'Tax Amount', 'Receipt']
    header_row = 3
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Data
    total_amount = 0
    total_tax = 0
    
    for row_idx, item in enumerate(items, header_row + 1):
        ws.cell(row=row_idx, column=1, value=str(item.get('date', '')))
        ws.cell(row=row_idx, column=2, value=item.get('vendor', ''))
        ws.cell(row=row_idx, column=3, value=item.get('description', ''))
        ws.cell(row=row_idx, column=4, value=item.get('category', ''))
        ws.cell(row=row_idx, column=5, value=item.get('amount', 0))
        ws.cell(row=row_idx, column=6, value=item.get('tax_amount', 0))
        ws.cell(row=row_idx, column=7, value='Yes' if item.get('receipt_available') else 'No')
        
        total_amount += item.get('amount', 0)
        total_tax += item.get('tax_amount', 0)
    
    # Totals
    total_row = header_row + len(items) + 2
    ws.cell(row=total_row, column=4, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=total_amount).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=total_tax).font = Font(bold=True)
    
    # Adjust columns
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
